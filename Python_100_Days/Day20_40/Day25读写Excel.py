# Day 25 读写Excel
# ==================== 基本信息 ====================
# ==================== 导入模块 ====================
import datetime
import openpyxl
import os
# ==================== 主 程 序 ====================
script_dir = os.path.dirname(os.path.abspath(__file__))
file_path = os.path.join(script_dir, 'res', '阿里巴巴2020年股票数据.xlsx')

wb = openpyxl.load_workbook(file_path)
print(wb.sheetnames)
sheet = wb.worksheets[0]
print(sheet.dimensions) # 获取单元格范围.两点确定一个长方形表格范围
print(sheet.max_row, sheet.max_column) # 获取行数和列数

print(sheet.cell(3, 3).value) # 获取指定单元格值
print(sheet['C3'].value)
print(sheet['G255'].value)

print(sheet['A2:C5'])         # 获取多个单元格的值
# 这种索引方式是用于直接访问工作表中的单元格区域
# 它不会直接返回单元格中的值，而是返回对应单元格对象本身

for row in range(2, sheet.max_row + 1):
    for col in 'ABCDEFG':
        value = sheet[f'{col}{row}'].value
        if type(value) == datetime.datetime:
            
            print(value.strtime('%Y年%m月%d日'), end='\t')
        elif type(value) == int:
            print(f'{value:<10d}', end='\t')
        elif type(value) == float:
            print(f'{value:.4f}', end='\t')
        else:
            print(value, end='\t')
    print()