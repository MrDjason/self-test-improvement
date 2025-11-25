# Day 25 读写Excel
# ==================== 基本信息 ====================
# ==================== 导入模块 ====================
import datetime
import openpyxl
import os
import random


# ==================== 主 程 序 ====================
script_dir = os.path.dirname(os.path.abspath(__file__))
file_path = os.path.join(script_dir, 'res', '阿里巴巴2020年股票数据.xlsx')

wb = openpyxl.load_workbook(file_path)
print(wb.sheetnames)
sheet = wb.worksheets[0]
print(sheet.dimensions) # 获取单元格范围.两点确定一个长方形表格范围
print(sheet.max_row, sheet.max_column) # 获取行数和列数

print(sheet.cell(3, 3).value) # 获取指定单元格值
print(sheet['C3'].value)
print(sheet['G255'].value)

print(sheet['A2:C5'])         # 获取多个单元格的值
# 这种索引方式是用于直接访问工作表中的单元格区域
# 它不会直接返回单元格中的值，而是返回对应单元格对象本身
'''读'''
for row in range(2, sheet.max_row + 1):
    for col in 'ABCDEFG':
        value = sheet[f'{col}{row}'].value
        if type(value) == datetime.datetime:
        # openpyxl 库在读取单元格数据时，会根据单元格的格式信息以及数据特征来判断是否为日期或时间类型。
            '''
            假如读到A2值
            读取到的值: 2019-12-31 00:00:00
            值的类型: <class 'datetime.datetime'>
            是否为 datetime.datetime 实例: True
            '''
            print(value.strftime('%Y年%m月%d日'), end='\t')
            # 按照指定格式打印这个日期时间 
            # strftime()是 datetime 对象的方法，用于将日期时间格式化为字符串
        elif type(value) == int:
            print(f'{value:<10d}', end='\t')
            # f'{value:<10d}'
            # d：表示按整数格式处理value
            # <：表示左对齐
            # 10：表示总宽度为10个字符
        elif type(value) == float:
            print(f'{value:.4f}', end='\t')
        else:
            print(value, end='\t')
    print()

'''写'''
wb = openpyxl.Workbook()
# 第一步 创建工作簿
sheet = wb.active # 获取工作簿中当前 “正在工作” 的工作表对象
sheet.name = '期末成绩'
# 第二步 添加工作表
titles = ('姓名', '语文', '数学', '英语')
for col, title in enumerate(titles):
    sheet.cell(1, col + 1, title) # 写入第一行标题
    # sheet.cell()的参数：行号，列号，值

names = ('关羽', '张飞', '赵云', '马超', '黄忠')
for row_index, name in enumerate(names):
    sheet.cell(row_index + 2, 1, name) # 从第二行到第六行第一列，写入姓名
    for col_index in range(2,5):
        sheet.cell(row_index + 2, col_index, random.randint(50, 100)) # random.randint(a, b) 返回一个[a, b]范围内的整数
    
wb.save(os.path.join(script_dir, 'res', '考试成绩表.xlsx'))

'''修改样式和公式计算'''
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side

alignment = Alignment(horizontal='center', vertical='center')
# 对齐方式：水平居中 垂直居中
side = Side(color='ff7f50', style='mediumDashed')
# 边框线条：珊瑚色 中等虚线


wb = openpyxl.load_workbook(os.path.join(script_dir, 'res', '考试成绩表.xlsx'))
sheet = wb.worksheets[0]

# 调整行高和列宽
sheet.row_dimensions[1].height = 30 # 设置第一行行高
sheet.column_dimensions['E'].width = 120 # 设置E列列宽

sheet['E1'] = '平均分' # 设置E1单元格的值为平均分

sheet.cell(1,5).font = Font(size=18, bold=True, color='ff1493', name='华文楷体')
# 设置E1单元格字体：18 加粗 深粉色 华文楷体
sheet.cell(1,5).alignment = alignment
# 设置对齐方式：水平居中 垂直居中
sheet.cell(1,5).border = Border(left=side, top=side, right=side, bottom=side)
# 设置单元格边框：左、上、右、下边框均使用前面定义的珊瑚色虚线
for i in range(2,7):
    sheet[f'E{i}'] = f'=average(B{i}:D{i})'
    # 设置Ei值为'=average(B{i}:D{i})'，到xlsx中计算B列到D列的平均分
    sheet.cell(i,5).font = Font(size=12, color='4169e1', italic=True)
    # 设置字体：12 宝蓝色 斜体
    sheet.cell(i,5).alignment = alignment
    # 设置对齐方式：水平居中 垂直居中
wb.save(os.path.join(script_dir, 'res', '考试成绩表.xlsx'))


'''生成统计图表'''
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference

wb = Workbook(write_only = True) # 创建一个只写的工作簿
sheet = wb.create_sheet()

rows = [
    ('类别', '销售A组', '销售B组'),
    ('手机', 40, 30),
    ('平板', 50, 60),
    ('笔记本', 80, 70),
    ('外围设备', 20, 10),
]

for row in rows:
    sheet.append(row) # sheet.append(row)会自动按顺序写入一行数据（从第一行开始）

chart = BarChart() # 创建一个柱状图对象
chart.type = 'col' #'col'表示垂直柱形图（默认）；若为'bar'则是水平条形图
chart.style = 10   # 设置图表样式：10是预设的样式编号，控制图表的颜色、线条等外观

chart.title = '销售统计图'      # 设置图表的标题（显示在图表上方）
chart.y_axis.title = '销量'   # 设置y轴（纵轴）的标题（描述坐标轴含义）
chart.x_axis.title = '商品类别' # 设置x轴（横轴）的标题（描述坐标轴含义）

data = Reference(sheet, min_col=2, min_row=1, max_row=5, max_col=3)

cats = Reference(sheet, min_col=1, min_row=2, max_row=5)

chart.add_data(data, titles_from_data=True)

chart.set_categories(cats)
chart.shape = 4

sheet.add_chart(chart, 'A10')

wb.save(os.path.join(script_dir, 'res', 'demo.xlsx'))